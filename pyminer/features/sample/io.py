import os
import sys
import logging
import numpy as np
import pandas as pd
import datetime

# 导入PyQt5模块
from PyQt5.QtGui import QCursor, QIcon, QPixmap
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

#导入功能组件
from pyminer.ui.data.data_import_text import Ui_Form as Import_Ui_Form
from pyminer.ui.data.data_import_excel import Ui_Form as ExcelImport_Ui_Form
from pyminer.ui.data.data_import_spss import Ui_Form as SPSSImport_Ui_Form
from pyminer.ui.data.data_import_sas import Ui_Form as SASImport_Ui_Form
from pyminer.ui.data.data_import_database import Ui_Form as ImportDatabase_Ui_Form

# 定义日志输出格式
logging.basicConfig(format="%(asctime)s %(name)s:%(levelname)s:%(message)s", datefmt="%Y-%m-%d %H:%M:%S",
                    level=logging.INFO)


class ImportForm(QDialog,Import_Ui_Form):
    """
    "导入"窗口
    """
    signal_data_change = pyqtSignal(str, dict, str, str, str, str, str)  # 自定义信号，用于传递文件路径

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.center()

        #QssTools.set_qss_to_obj(ui_dir + "/source/qss/patata.qss", self)

        self.file_path = ''
        self.current_dataset_name = ""
        self.current_dataset = pd.DataFrame()

        # self.import_file_path_init()

        # 导入窗口的相关事件
        # 在"导入"窗口，打开选择文件
        self.pushButton_choosefile.clicked.connect(self.openFile)
        # 展示数据
        self.checkBox_ifColumns.stateChanged.connect(self.import_dateset_reload)
        self.comboBox_separator.currentTextChanged.connect(self.import_dateset_reload)
        self.comboBox_encode.currentTextChanged.connect(self.import_dateset_reload)
        self.lineEdit_filePath.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_passHead.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_datasetName.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_limitRow.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_missValue.textChanged.connect(self.import_dateset_reload)

        # 更新数据
        self.pushButton_ok.clicked.connect(self.import_send_dataset)
        self.pushButton_cancel.clicked.connect(self.close)
        # 帮助
        # self.pushButton_help.clicked.connect(self.accept_signal)

    def file_path_init(self):
        print("开始更新数据")
        print("开始修改文件路径")
        self.file_path = self.file_path
        if len(self.file_path) != 0:
            self.lineEdit_filePath.setText(self.file_path)

            self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
            self.lineEdit_datasetName.setText(self.current_dataset_name)
            logging.info(
                "加载成功file_path{}，datasetName：{}".format(self.file_path, self.current_dataset_name))

            if len(self.file_path) > 0:
                if self.checkBox_ifColumns.isChecked():
                    header = 0
                else:
                    header = None
                # 仅预览前100条数据
                if self.lineEdit_limitRow.text() == "全部":
                    nrows_preview = 100
                elif int(self.lineEdit_limitRow.text()) <= 100:
                    nrows_preview = int(self.lineEdit_limitRow.text())
                else:
                    nrows_preview = 100

                if self.lineEdit_limitRow.text() == "全部":
                    nrows = 100000000
                else:
                    nrows = int(self.lineEdit_limitRow.text())

                encoding = self.comboBox_encode.currentText()
                skiprows = int(self.lineEdit_passHead.text())
                sep = self.comboBox_separator.currentText()

                if self.lineEdit_missValue.text() != "默认":
                    na_values = self.lineEdit_missValue.text()
                else:
                    na_values = None

                logging.info("file_path：{}，header：{}，skiprows：{}，nrows：{}，na_values:{}".format(
                    self.file_path, header, skiprows, nrows_preview, na_values))

                self.current_dataset = pd.read_csv(self.file_path, engine="python", sep=sep,
                                                     encoding=encoding,
                                                     header=header,
                                                     skiprows=skiprows, nrows=nrows, na_values=na_values)

                if len(self.current_dataset) == 0:
                    self.tableWidget_previewData.clear()
                    logging.info("当前有效数据为空")
                else:
                    self.import_dateset_preview()
                    logging.info("数据导入成功")
            else:
                logging.info("请选择数据集")

    def import_send_dataset(self):
        logging.info("发射导入数据信号")
        if len(self.current_dataset) > 0:
            create_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据创建时间
            update_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据更新时间
            path = self.file_path
            print("path:", path)
            file_size=str(os.path.getsize(path))
            print("file_size:",file_size)
            remarks = ''
            self.signal_data_change.emit(self.current_dataset_name, self.current_dataset.to_dict(), path,
                                         create_time, update_time, remarks, file_size)  # 发射信号
            logging.info("导入数据信号已发射")
            self.close()
        else:
            logging.info("导入数据信号发射失败")
            self.close()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def keyPressEvent(self, e):
        """
        按键盘Escape退出当前窗口
        @param e:
        """
        if e.key() == Qt.Key_Escape:
            self.close()

    def openFile(self):
        """
        选择文件
        """
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', '文本文件(*.csv *.txt *.tsv)')
        self.file_path = openfile_name[0]
        self.lineEdit_filePath.setText(self.file_path)

        self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
        self.lineEdit_datasetName.setText(self.current_dataset_name)

    def import_dateset_reload(self):
        """
        刷新导入的数据
        """
        header = 0
        nrows_preview = 100
        sep = ','
        skiprows = 0

        if len(self.file_path) > 0:
            if self.checkBox_ifColumns.isChecked():
                header = 'infer'
            else:
                header = None
            # 仅预览前100条数据
            if self.lineEdit_limitRow.text() == "全部":
                nrows_preview = 100
            elif int(self.lineEdit_limitRow.text()) <= 100:
                nrows_preview = int(self.lineEdit_limitRow.text())
            else:
                nrows_preview = 100

            if self.lineEdit_limitRow.text() == "全部":
                nrows = 100000000
            else:
                nrows = int(self.lineEdit_limitRow.text())

            encoding = self.comboBox_encode.currentText()
            sep = self.comboBox_separator.currentText()
            skiprows = int(self.lineEdit_passHead.text())

            if self.lineEdit_limitRow.text() != "默认":
                na_values = self.lineEdit_missValue.text()

            self.current_dataset = pd.read_csv(self.file_path, engine="python", sep=sep, encoding=encoding,
                                                 header=header,
                                                 skiprows=skiprows, nrows=nrows, na_values=na_values)

            if len(self.current_dataset) == 0:
                self.tableWidget_previewData.clear()
                logging.info("当前有效数据为空")
            else:
                self.import_dateset_preview()
                logging.info("数据导入成功")

    def import_dateset_preview(self):
        """
        刷新预览数据
        """
        if len(self.current_dataset) > 0:
            input_table_rows = self.current_dataset.head(100).shape[0]
            input_table_colunms = self.current_dataset.shape[1]
            input_table_header = self.current_dataset.columns.values.tolist()
            self.tableWidget_previewData.setColumnCount(input_table_colunms)
            self.tableWidget_previewData.setRowCount(input_table_rows)

            # 设置数据预览窗口的标题行
            table_header = []
            i = 1
            while i <= len(self.current_dataset.columns):
                table_header.append("C" + str(i))
                i += 1

            if self.checkBox_ifColumns.isChecked():
                self.tableWidget_previewData.setHorizontalHeaderLabels(input_table_header)
            else:
                self.tableWidget_previewData.setHorizontalHeaderLabels(table_header)
                self.current_dataset.columns = table_header
            # 数据预览窗口
            for i in range(input_table_rows):
                input_table_rows_values = self.current_dataset.iloc[[i]]
                input_table_rows_values_array = np.array(input_table_rows_values)
                input_table_rows_values_list = input_table_rows_values_array.tolist()[0]
                for j in range(input_table_colunms):
                    input_table_items_list = input_table_rows_values_list[j]

                    input_table_items = str(input_table_items_list)
                    newItem = QTableWidgetItem(input_table_items)
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_previewData.setItem(i, j, newItem)


class ImportDatabase(QDialog,ImportDatabase_Ui_Form):
    """
    "导入"窗口
    """
    signal_data_change = pyqtSignal(str, dict, str, str, str, str, str)  # 自定义信号，用于传递文件路径

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setupUi(self)
        self.center()

        self.current_dataset_name = ''  # 当前数据集名称
        self.current_dataset = pd.DataFrame()  # 修改后的数据集
        self.all_dataset=''  #当前已导入的全部数据
        self.file_path='' #当前数据路径

        #QssTools.set_qss_to_obj(ui_dir + "/source/qss/patata.qss", self)
        self.label_test.setHidden(True)
        self.lineEdit_passwd.setEchoMode(QtWidgets.QLineEdit.Password)

        #事件
        self.pushButton_cancel.clicked.connect(self.close)
        self.pushButton_test.clicked.connect(self.database_test)
        self.pushButton_ok.clicked.connect(self.import_send_dataset)

    def keyPressEvent(self, e):
        """
        按键盘Escape退出当前窗口
        """
        if e.key() == Qt.Key_Escape:
            self.close()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def import_send_dataset(self):
        logging.info("发射导入数据信号")
        if len(self.current_dataset) > 0:
            self.signal_data_change.emit(self.current_dataset_name, self.current_dataset.to_dict())  # 发射信号
            logging.info("导入数据信号已发射")
            self.close()
        else:
            logging.info("导入数据信号发射失败")
            self.close()


    def database_test(self):
        import pymysql
        host=self.lineEdit_host.text()
        user=self.lineEdit_user.text()
        passwd=self.lineEdit_passwd.text()
        db=self.lineEdit_db.text()
        port=self.spinBox_port.value()
        charset = 'utf8'
        table=self.lineEdit_table.text()
        sql='select * from '+db+'.'+table;
        print(sql)
        if len(db)==0:
            QMessageBox.information(self, '注意', '数据库名不能为空', QMessageBox.Ok , QMessageBox.Ok)
        elif len(table)==0:
            QMessageBox.information(self, '注意', '表名不能为空', QMessageBox.Ok, QMessageBox.Ok)
        else:
            try:
                conn = pymysql.connect(host=host, user=user, passwd=passwd, db=db, port=port, charset=charset)
                cur = conn.cursor()
                cur.execute(sql)
                conn.close()
                print('连接成功')
                self.label_test.setHidden(False)
                self.label_test.setText('连接成功')
                self.label_test.setStyleSheet('color: blue;')

            except Exception as Error:
                print('连接失败:' + str(Error))
                self.label_test.setHidden(False)
                self.label_test.setText('连接失败:' + str(Error))
                self.label_test.setStyleSheet('color: rgb(255, 0, 0);')

    def database_connect(self):
        import pymysql
        host=self.lineEdit_host.text()
        user=self.lineEdit_user.text()
        passwd=self.lineEdit_passwd.text()
        db=self.lineEdit_db.text()
        port=self.spinBox_port.value()
        charset = 'utf8'
        table=self.lineEdit_table.text()
        self.current_dataset_name=table
        self.file_path=db+'.'+table
        sql='select * from '+self.file_path;
        print(sql)
        if len(db)==0:
            QMessageBox.information(self, '注意', '数据库名不能为空', QMessageBox.Ok , QMessageBox.Ok)
        elif len(table)==0:
            QMessageBox.information(self, '注意', '表名不能为空', QMessageBox.Ok, QMessageBox.Ok)
        else:
            try:
                conn = pymysql.connect(host=host, user=user, passwd=passwd, db=db, port=port, charset=charset)
                df = pd.read_sql(sql, con = conn)
                return df
                conn.close()
                print('连接成功')
                self.label_test.setHidden(False)
                self.label_test.setText('连接成功')
                self.label_test.setStyleSheet('color: blue;')

            except Exception as Error:
                print('连接失败:' + str(Error))
                self.label_test.setHidden(False)
                self.label_test.setText('连接失败:' + str(Error))
                self.label_test.setStyleSheet('color: rgb(255, 0, 0);')

    def import_send_dataset(self):
        self.current_dataset = self.database_connect()
        logging.info("发射导入数据信号")
        if len(self.current_dataset) > 0:
            create_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据创建时间
            update_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据更新时间
            path = self.file_path
            file_size = ''
            remarks = ''
            self.signal_data_change.emit(self.current_dataset_name, self.current_dataset.to_dict(), path,
                                         create_time, update_time, remarks, file_size)  # 发射信号
            self.close()
        else:
            logging.info("导入数据信号发射失败")
            self.close()


class ImportExcelForm(QDialog,ExcelImport_Ui_Form):
    """
    打开"从excel导入"窗口
    """
    signal_data_change = pyqtSignal(str, dict, str, str, str, str, str)  # 自定义信号，用于传递文件路径

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.center()

        self.file_path = ''  #文件路径
        self.current_dataset_name = "" #当前数据集名称
        self.current_dataset = pd.DataFrame() #当前数据集

        # 导入窗口的相关事件
        # 在"导入"窗口，打开选择文件
        self.pushButton_choosefile.clicked.connect(self.openFile)
        # 展示数据
        self.checkBox_ifColumns.stateChanged.connect(self.import_dateset_reload)
        self.comboBox_sheet.currentTextChanged.connect(self.import_dateset_reload)
        self.comboBox_encode.currentTextChanged.connect(self.import_dateset_reload)
        self.lineEdit_filePath.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_passHead.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_datasetName.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_limitRow.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_missValue.textChanged.connect(self.import_dateset_reload)

        # 更新数据
        self.pushButton_ok.clicked.connect(self.import_send_dataset)
        self.pushButton_cancel.clicked.connect(self.close)

    def keyPressEvent(self, e):
        """
        按键盘Escape退出当前窗口
        @param e:
        """
        if e.key() == Qt.Key_Escape:
            self.close()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def import_send_dataset(self):
        logging.info("发射导入数据信号")
        if len(self.current_dataset) > 0:
            create_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据创建时间
            update_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据更新时间
            path = self.file_path
            file_size = str(os.path.getsize(path))
            remarks = ''
            self.signal_data_change.emit(self.current_dataset_name, self.current_dataset.to_dict(), path,
                                         create_time, update_time, remarks, file_size)  # 发射信号
            self.close()
        else:
            logging.info("导入数据信号发射失败")
            self.close()

    def file_path_init(self, file_path):
        """
        #初始化excel文件路径、sheet名单,以便指定需要导入的sheet
        """
        self.file_path = file_path
        print("self.file_path:", self.file_path)
        if len(self.file_path) != 0:
            self.lineEdit_filePath.setText(self.file_path)

            import openpyxl
            wb = openpyxl.load_workbook(self.file_path)

            # 获取excel 工作簿中所有的sheet
            sheets = wb.sheetnames
            self.comboBox_sheet.clear()
            for s in sheets:
                self.comboBox_sheet.addItem(s)

            self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
            self.lineEdit_datasetName.setText(self.current_dataset_name)
            logging.info(
                "加载成功file_path{}，datasetName：{}".format(self.file_path, self.current_dataset_name))

            if len(self.file_path) > 0:
                if self.checkBox_ifColumns.isChecked():
                    header = 0
                else:
                    header = None
                # 仅预览前100条数据
                if self.lineEdit_limitRow.text() == "全部":
                    nrows_preview = 100
                elif int(self.lineEdit_limitRow.text()) <= 100:
                    nrows_preview = int(self.lineEdit_limitRow.text())
                else:
                    nrows_preview = 100

                if self.lineEdit_limitRow.text() == "全部":
                    nrows = 100000000
                else:
                    nrows = int(self.lineEdit_limitRow.text())

                skiprows = int(self.lineEdit_passHead.text())
                sheet = self.comboBox_sheet.currentText()

                if self.lineEdit_missValue.text() != "默认":
                    na_values = self.lineEdit_missValue.text()
                else:
                    na_values = None

                logging.info("file_path：{}，sheet_name：{}，header：{}，skiprows：{}，nrows：{}，na_values:{}".format(
                    self.file_path, sheet, header, skiprows, nrows_preview, na_values))

                self.current_dataset = pd.read_excel(self.file_path, sheet_name=sheet,
                                                       header=header,
                                                       skiprows=skiprows, nrows=nrows, na_values=na_values)

                if len(self.current_dataset) == 0:
                    self.tableWidget_previewData.clear()
                    logging.info("当前sheet页有效数据为空")
                else:
                    self.import_dateset_preview()
                    logging.info("数据导入成功")
            else:
                logging.info("请选择数据集")

    def openFile(self):
        """
        选择文件
        """
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'EXCEL文件(*.xls *.xlsx *.xlsm *.xltx *.xltm)')
        logging.info(openfile_name)

        self.file_path = openfile_name[0]
        self.lineEdit_filePath.setText(self.file_path)

        # 获取excel 工作簿中所有的sheet
        import openpyxl
        wb = openpyxl.load_workbook(self.file_path)
        sheets = wb.sheetnames
        self.comboBox_sheet.clear()
        for s in sheets:
            self.comboBox_sheet.addItem(s)

        self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
        self.lineEdit_datasetName.setText(self.current_dataset_name)
        logging.info("加载成功file_path{}，datasetName：{}".format(openfile_name, self.current_dataset_name))
        self.import_dateset_reload()

    def import_dateset_reload(self):
        """
        刷新导入的数据
        """
        header = 0
        nrows_preview = 100
        sep = ','
        skiprows = 0
        # datasetName 为当前已选文件对应的excel数据路径

        if len(self.file_path) > 0:
            if self.checkBox_ifColumns.isChecked():
                header = 0
            else:
                header = None
            # 仅预览前100条数据
            if self.lineEdit_limitRow.text() == "全部":
                nrows_preview = 100
            elif int(self.lineEdit_limitRow.text()) <= 100:
                nrows_preview = int(self.lineEdit_limitRow.text())
            else:
                nrows_preview = 100

            if self.lineEdit_limitRow.text() == "全部":
                nrows = 100000000
            else:
                nrows = int(self.lineEdit_limitRow.text())

            encoding = self.comboBox_encode.currentText()
            skiprows = int(self.lineEdit_passHead.text())
            sheet = self.comboBox_sheet.currentText()

            if self.lineEdit_missValue.text() != "默认":
                na_values = self.lineEdit_missValue.text()
            else:
                na_values = None

            logging.info("file_path：{}，sheet_name：{}，header：{}，skiprows：{}，nrows：{}，na_values:{}".format(
                self.file_path, sheet, header, skiprows, nrows_preview, na_values))
            if len(sheet)>0:
                self.current_dataset = pd.read_excel(self.file_path, sheet_name=sheet,
                                                     header=header,
                                                     skiprows=skiprows, nrows=nrows, na_values=na_values)

                if len(self.current_dataset) == 0:
                    self.tableWidget_previewData.clear()
                    logging.info("当前sheet页有效数据为空")
                else:
                    self.import_dateset_preview()
                    logging.info("数据导入成功")

    def import_dateset_preview(self):
        """
        刷新预览数据
        """
        if len(self.current_dataset) > 0:
            input_table_rows = self.current_dataset.head(100).shape[0]  # 预览前100行
            input_table_colunms = self.current_dataset.shape[1]
            input_table_header = self.current_dataset.columns.values.tolist()
            self.tableWidget_previewData.setColumnCount(input_table_colunms)
            self.tableWidget_previewData.setRowCount(input_table_rows)

            # 设置数据预览窗口的标题行
            table_header = []
            i = 1
            while i <= len(self.current_dataset.columns):
                table_header.append("C" + str(i))
                i += 1

            if self.checkBox_ifColumns.isChecked():
                self.tableWidget_previewData.setHorizontalHeaderLabels(input_table_header)
            else:
                self.tableWidget_previewData.setHorizontalHeaderLabels(table_header)
                self.current_dataset.columns = table_header
            # 数据预览窗口
            for i in range(input_table_rows):
                input_table_rows_values = self.current_dataset.iloc[[i]]
                input_table_rows_values_array = np.array(input_table_rows_values)
                input_table_rows_values_list = input_table_rows_values_array.tolist()[0]
                for j in range(input_table_colunms):
                    input_table_items_list = input_table_rows_values_list[j]

                    input_table_items = str(input_table_items_list)
                    newItem = QTableWidgetItem(input_table_items)
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_previewData.setItem(i, j, newItem)



class ImportSpssForm(QDialog,SPSSImport_Ui_Form):
    """
    打开"从spss导入"窗口
    """
    signal_data_change = pyqtSignal(str, dict, str, str, str, str, str)  # 自定义信号，用于传递文件路径

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.center()

        #QssTools.set_qss_to_obj(ui_dir + "/source/qss/patata.qss", self)

        self.file_path = ''
        self.current_dataset_name = ""
        self.current_dataset = pd.DataFrame()

        self.all_dataset = dict()

        # 导入窗口的相关事件
        # 在"导入"窗口，打开选择文件
        self.pushButton_choosefile.clicked.connect(self.openFile)
        # 展示数据
        self.checkBox_ifColumns.stateChanged['int'].connect(self.import_dateset_reload)
        self.comboBox_encode.currentTextChanged['QString'].connect(self.import_dateset_reload)
        self.lineEdit_filePath.textChanged['QString'].connect(self.import_dateset_reload)
        self.lineEdit_passHead.textChanged['QString'].connect(self.import_dateset_reload)
        self.lineEdit_datasetName.textChanged['QString'].connect(self.import_dateset_reload)
        self.lineEdit_limitRow.textChanged['QString'].connect(self.import_dateset_reload)

        # 更新数据
        self.pushButton_ok.clicked.connect(self.import_send_dataset)
        self.pushButton_cancel.clicked.connect(self.close)

    def keyPressEvent(self, e):
        """
        按键盘Escape退出当前窗口
        @param e:
        """
        if e.key() == Qt.Key_Escape:
            self.close()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def import_send_dataset(self):
        logging.info("发射导入数据信号")
        if len(self.current_dataset) > 0:
            create_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据创建时间
            update_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据更新时间
            path = self.file_path
            file_size = self.all_dataset.get("")
            remarks = ''
            self.signal_data_change.emit(self.current_dataset_name, self.current_dataset.to_dict(), path,
                                         create_time, update_time, remarks, file_size)  # 发射信号
            self.close()
        else:
            logging.info("导入数据信号发射失败")
            self.close()

    def file_path_init(self, file_path):
        """
        #初始化spss文件路径
        """
        self.file_path = file_path
        if len(self.file_path) != 0:
            self.lineEdit_filePath.setText(self.file_path)
            self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
            self.lineEdit_datasetName.setText(self.current_dataset_name)
            logging.info(
                "加载成功file_path{}，datasetName：{}".format(self.file_path, self.current_dataset_name))

            if len(self.file_path) > 0:
                if self.checkBox_ifColumns.isChecked():
                    header = 0
                else:
                    header = None
                # 仅预览前100条数据
                if self.lineEdit_limitRow.text() == "全部":
                    nrows_preview = 100
                elif int(self.lineEdit_limitRow.text()) <= 100:
                    nrows_preview = int(self.lineEdit_limitRow.text())
                else:
                    nrows_preview = 100

                if self.lineEdit_limitRow.text() == "全部":
                    nrows = 100000000
                else:
                    nrows = int(self.lineEdit_limitRow.text())

                encoding = self.comboBox_encode.currentText()
                skiprows = int(self.lineEdit_passHead.text())

                logging.info("file_path：{}，header：{}，skiprows：{}，nrows：{}".format(
                    self.file_path, header, skiprows, nrows_preview))

                self.current_dataset = pd.read_spss(self.file_path)

                if len(self.current_dataset) == 0:
                    self.tableWidget_previewData.clear()
                    logging.info("当前有效数据为空")
                else:
                    self.import_dateset_preview()
                    logging.info("数据导入成功")
            else:
                print("请选择数据集")

    def openFile(self):
        """
        选择文件
        """
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'SPSS文件(*.sav)')
        logging.info(openfile_name)

        self.file_path = openfile_name[0]
        self.lineEdit_filePath.setText(self.file_path)

        self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
        self.lineEdit_datasetName.setText(self.current_dataset_name)
        logging.info("加载成功file_path{}，datasetName：{}".format(openfile_name, self.current_dataset_name))
        self.import_dateset_reload()

    def get_dateset_name(self):
        return self.current_dataset_name

    def import_dateset_reload(self):
        """
        刷新导入的数据
        """
        header = 0
        nrows_preview = 100
        sep = ','
        skiprows = 0

        if len(self.file_path) > 0:
            if self.checkBox_ifColumns.isChecked():
                header = 0
            else:
                header = None
            # 仅预览前100条数据
            if self.lineEdit_limitRow.text() == "全部":
                nrows_preview = 100
            elif int(self.lineEdit_limitRow.text()) <= 100:
                nrows_preview = int(self.lineEdit_limitRow.text())
            else:
                nrows_preview = 100

            if self.lineEdit_limitRow.text() == "全部":
                nrows = 100000000
            else:
                nrows = int(self.lineEdit_limitRow.text())

            encoding = self.comboBox_encode.currentText()
            skiprows = int(self.lineEdit_passHead.text())

            logging.info("file_path：{}，header：{}，skiprows：{}，nrows：{}".format(
                self.file_path, header, skiprows, nrows_preview))

            self.current_dataset = pd.read_spss(self.file_path)

            if len(self.current_dataset) == 0:
                self.tableWidget_previewData.clear()
                logging.info("当前有效数据为空")
            else:
                self.import_dateset_preview()
                logging.info("数据导入成功")

    def import_dateset_preview(self):
        """
        刷新预览数据
        """
        if len(self.current_dataset) > 0:
            input_table_rows = self.current_dataset.head(100).shape[0]
            input_table_colunms = self.current_dataset.shape[1]
            input_table_header = self.current_dataset.columns.values.tolist()
            self.tableWidget_previewData.setColumnCount(input_table_colunms)
            self.tableWidget_previewData.setRowCount(input_table_rows)

            # 设置数据预览窗口的标题行
            table_header = []
            i = 1
            while i <= len(self.current_dataset.columns):
                table_header.append("C" + str(i))
                i += 1

            if self.checkBox_ifColumns.isChecked():
                self.tableWidget_previewData.setHorizontalHeaderLabels(input_table_header)
            else:
                self.tableWidget_previewData.setHorizontalHeaderLabels(table_header)
                self.current_dataset.columns = table_header
            # 数据预览窗口
            for i in range(input_table_rows):
                input_table_rows_values = self.current_dataset.iloc[[i]]
                input_table_rows_values_array = np.array(input_table_rows_values)
                input_table_rows_values_list = input_table_rows_values_array.tolist()[0]
                for j in range(input_table_colunms):
                    input_table_items_list = input_table_rows_values_list[j]

                    input_table_items = str(input_table_items_list)
                    newItem = QTableWidgetItem(input_table_items)
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_previewData.setItem(i, j, newItem)

    def get_current_dataset(self):
        return self.current_dataset

    def get_current_columns(self):
        return self.current_dataset.columns


class ImportSasForm(QDialog,SASImport_Ui_Form):
    """
    打开"从sas导入"窗口
    """
    signal_data_change = pyqtSignal(str, dict, str, str, str, str, str)  # 自定义信号，用于传递文件路径

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.center()

        self.file_path = ''
        self.current_dataset_name = ""
        self.current_dataset = pd.DataFrame()

        # 导入窗口的相关事件
        # 在"导入"窗口，打开选择文件
        self.pushButton_choosefile.clicked.connect(self.openFile)
        # 展示数据
        self.checkBox_ifColumns.stateChanged.connect(self.import_dateset_reload)

        self.comboBox_encode.currentTextChanged.connect(self.import_dateset_reload)
        self.lineEdit_filePath.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_passHead.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_datasetName.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_limitRow.textChanged.connect(self.import_dateset_reload)
        self.lineEdit_missValue.textChanged.connect(self.import_dateset_reload)

        # 更新数据
        self.pushButton_ok.clicked.connect(self.import_send_dataset)
        self.pushButton_cancel.clicked.connect(self.close)

    def keyPressEvent(self, e):
        """
        按键盘Escape退出当前窗口
        @param e:
        """
        if e.key() == Qt.Key_Escape:
            self.close()

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def import_send_dataset(self):
        logging.info("发射导入数据信号")
        if len(self.current_dataset) > 0:
            create_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据创建时间
            update_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 数据更新时间
            path = self.file_path
            file_size = str(os.path.getsize(path))
            remarks = ''
            self.signal_data_change.emit(self.current_dataset_name, self.current_dataset.to_dict(), path,
                                         create_time, update_time, remarks, file_size)  # 发射信号
            self.close()
        else:
            logging.info("导入数据信号发射失败")
            self.close()

    def file_path_init(self, file_path):
        """
        #初始化sas文件路径
        """
        self.file_path = file_path
        if len(self.file_path) != 0:
            self.lineEdit_filePath.setText(self.file_path)

            self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
            self.lineEdit_datasetName.setText(self.current_dataset_name)
            logging.info(
                "加载成功file_path{}，datasetName：{}".format(self.file_path, self.current_dataset_name))

            if len(self.file_path) > 0:
                if self.checkBox_ifColumns.isChecked():
                    header = 0
                else:
                    header = None
                # 仅预览前100条数据
                if self.lineEdit_limitRow.text() == "全部":
                    nrows_preview = 100
                elif int(self.lineEdit_limitRow.text()) <= 100:
                    nrows_preview = int(self.lineEdit_limitRow.text())
                else:
                    nrows_preview = 100

                if self.lineEdit_limitRow.text() == "全部":
                    nrows = 100000000
                else:
                    nrows = int(self.lineEdit_limitRow.text())

                encoding = self.comboBox_encode.currentText()
                skiprows = int(self.lineEdit_passHead.text())

                logging.info("file_path：{}，header：{}，skiprows：{}，nrows：{}".format(
                    self.file_path, header, skiprows, nrows_preview))

                self.current_dataset = pd.read_sas(self.file_path, format='sas7bdat', encoding=encoding)

                if len(self.current_dataset) == 0:
                    self.tableWidget_previewData.clear()
                    logging.info("当前有效数据为空")
                else:
                    self.import_dateset_preview()
                    logging.info("数据导入成功")
            else:
                print("请选择数据集")

    def openFile(self):
        """
        选择文件
        """
        openfile_name = QFileDialog.getOpenFileName(self, '选择文件', '', 'SAS文件(*.sas7bdat)')
        logging.info(openfile_name)

        self.file_path = openfile_name[0]
        self.lineEdit_filePath.setText(self.file_path)

        self.current_dataset_name = os.path.split(self.lineEdit_filePath.text())[1]
        self.lineEdit_datasetName.setText(self.current_dataset_name)
        logging.info("加载成功file_path{}，datasetName：{}".format(openfile_name, self.current_dataset_name))
        self.import_dateset_reload()

    def import_dateset_reload(self):
        """
        刷新导入的数据
        """
        header = 0
        nrows_preview = 100
        sep = ','
        skiprows = 0

        if len(self.file_path) > 0:
            if self.checkBox_ifColumns.isChecked():
                header = 0
            else:
                header = None
            # 仅预览前100条数据
            if self.lineEdit_limitRow.text() == "全部":
                nrows_preview = 100
            elif int(self.lineEdit_limitRow.text()) <= 100:
                nrows_preview = int(self.lineEdit_limitRow.text())
            else:
                nrows_preview = 100

            if self.lineEdit_limitRow.text() == "全部":
                nrows = 100000000
            else:
                nrows = int(self.lineEdit_limitRow.text())

            encoding = self.comboBox_encode.currentText()
            skiprows = int(self.lineEdit_passHead.text())

            logging.info("file_path：{}，header：{}，skiprows：{}，nrows：{}".format(
                self.file_path, header, skiprows, nrows_preview))

            self.current_dataset = pd.read_sas(self.file_path, format='sas7bdat', encoding=encoding)

            if len(self.current_dataset) == 0:
                self.tableWidget_previewData.clear()
                logging.info("当前有效数据为空")
            else:
                self.import_dateset_preview()
                logging.info("数据导入成功")

    def import_dateset_preview(self):
        """
        刷新预览数据
        """
        if len(self.current_dataset) > 0:
            input_table_rows = self.current_dataset.head(100).shape[0]
            input_table_colunms = self.current_dataset.shape[1]
            input_table_header = self.current_dataset.columns.values.tolist()
            self.tableWidget_previewData.setColumnCount(input_table_colunms)
            self.tableWidget_previewData.setRowCount(input_table_rows)

            # 设置数据预览窗口的标题行
            table_header = []
            i = 1
            while i <= len(self.current_dataset.columns):
                table_header.append("C" + str(i))
                i += 1

            if self.checkBox_ifColumns.isChecked():
                self.tableWidget_previewData.setHorizontalHeaderLabels(input_table_header)
            else:
                self.tableWidget_previewData.setHorizontalHeaderLabels(table_header)
                self.current_dataset.columns = table_header
            # 数据预览窗口
            for i in range(input_table_rows):
                input_table_rows_values = self.current_dataset.iloc[[i]]
                input_table_rows_values_array = np.array(input_table_rows_values)
                input_table_rows_values_list = input_table_rows_values_array.tolist()[0]
                for j in range(input_table_colunms):
                    input_table_items_list = input_table_rows_values_list[j]

                    input_table_items = str(input_table_items_list)
                    newItem = QTableWidgetItem(input_table_items)
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_previewData.setItem(i, j, newItem)

    def get_current_dataset(self):
        return self.current_dataset

    def get_current_columns(self):
        return self.current_dataset.columns


# ====================================窗体测试程序============================
if __name__ == '__main__':
    app = QApplication(sys.argv)
    form = ImportForm()
    form.show()
    sys.exit(app.exec_())
